# excel_utils.py

from openpyxl import Workbook, load_workbook

def append_row_to_excel(path, row, header=None):
    if path.exists():
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        if header:
            ws.append(header)

    ws.append(row)
    wb.save(path)
