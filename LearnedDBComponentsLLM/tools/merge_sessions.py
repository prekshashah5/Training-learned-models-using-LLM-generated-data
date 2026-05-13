import os
import sys
from pathlib import Path
from dotenv import load_dotenv
from datetime import datetime
import json

# Add to path so we can import utils
sys.path.append(os.path.dirname(os.path.abspath(__file__)))
from utils.io_utils import read_json_file, save_output_metadata

def merge_sessions(target_limit=None):
    load_dotenv()
    
    # Resolve output folder
    if Path("output").exists() and Path("output").is_dir():
        OUTPUT_FOLDER = Path("output")
    else:
        OUTPUT_FOLDER = Path(os.getenv("OUTPUT_FOLDER", "../output"))
    
    if target_limit is None:
        target_limit = int(os.getenv("TOTAL_QUERIES", 30))
        
    TEMPERATURE = float(os.getenv("TEMPERATURE", 0.8))
    
    print(f"\n[info] Target limit per model: {target_limit} queries")
    
    if not OUTPUT_FOLDER.exists():
        print(f"[error] Output folder {OUTPUT_FOLDER} does not exist.")
        return
        
    # Get all session directories (excluding already merged ones)
    session_dirs = sorted(
        [d for d in OUTPUT_FOLDER.iterdir() if d.is_dir() and d.name.startswith("session_") and "merged" not in d.name], 
        key=lambda x: x.name
    )
    
    if not session_dirs:
        print("[warn] No standard sessions found to merge.")
        return
        
    print(f"[info] Scanning {len(session_dirs)} previous session folders...")
    
    from openpyxl import load_workbook
    
    merged_queries = {}
    merged_times = {}
    seen_sqls = {}
    
    for session_dir in session_dirs:
        # Load the metadata map for this session
        session_excel = session_dir / "runs.xlsx"
        session_run_meta = {}
        if session_excel.exists():
            try:
                wb = load_workbook(session_excel)
                ws = wb.active
                headers = [cell.value for cell in ws[1][:10]] # Limit header search
                run_id_idx = headers.index("Run_ID") if "Run_ID" in headers else -1
                gen_time_idx = headers.index("Generation_Time_s") if "Generation_Time_s" in headers else -1
                num_q_idx = headers.index("Num_Queries") if "Num_Queries" in headers else -1
                
                if run_id_idx >= 0 and gen_time_idx >= 0:
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        r_id = row[run_id_idx]
                        if r_id:
                            q_count = row[num_q_idx] if num_q_idx >= 0 else 0
                            g_time = row[gen_time_idx] if gen_time_idx >= 0 else 0
                            session_run_meta[r_id] = {"queries": q_count, "time": g_time}
            except Exception as e:
                print(f"[warn] Failed to read {session_excel}: {e}")

        for run_dir in session_dir.iterdir():
            if not run_dir.is_dir() or "_run_" not in run_dir.name:
                continue
                
            model_name_safe = run_dir.name.split("_run_")[0]
            model_name = model_name_safe.replace("_", ":")
            
            if model_name not in merged_queries:
                merged_queries[model_name] = []
                merged_times[model_name] = 0.0
                seen_sqls[model_name] = set()
                
            if len(merged_queries[model_name]) >= target_limit:
                continue # Already reached limit for this model
                
            queries_path = run_dir / "queries.jsonl"
            if not queries_path.exists():
                continue
                
            # Attempt to load queries safely
            try:
                queries = read_json_file(str(queries_path))
            except Exception as e:
                print(f"[warn] Failed to read {queries_path}: {e}")
                continue
                
            if not queries:
                continue
                
            added_this_run = 0
            for q in queries:
                if len(merged_queries[model_name]) >= target_limit:
                    break
                    
                # De-duplicate by exact SQL string
                sql = q.get("sql", "").strip()
                if sql and sql not in seen_sqls[model_name]:
                    seen_sqls[model_name].add(sql)
                    merged_queries[model_name].append(q)
                    added_this_run += 1
                    
            if added_this_run > 0:
                run_id = run_dir.name
                if run_id in session_run_meta:
                    meta = session_run_meta[run_id]
                    if meta["queries"] and meta["queries"] > 0 and meta["time"]:
                        # Distribute generation time proportionally relative to how many local queries were kept
                        time_added = (added_this_run / meta["queries"]) * float(meta["time"])
                        merged_times[model_name] += time_added
                        
    # Check if we collected anything
    if not sum(len(q) for q in merged_queries.values()):
        print("[warn] No valid queries were found to merge.")
        return
        
    # Create the new merged session directory
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    merged_session_id = f"session_merged_{timestamp}"
    merged_session_dir = OUTPUT_FOLDER / merged_session_id
    merged_session_dir.mkdir(parents=True, exist_ok=True)
    
    print(f"\n[info] Created merged session: {merged_session_dir}")
    
    for model_name, queries in merged_queries.items():
        if not queries:
            continue
            
        print(f"  - {model_name}: {len(queries)} unique queries merged.")
        
        safe_model = model_name.replace("/", "_").replace(" ", "_").replace(":", "_")
        run_id = f"{safe_model}_run_merged_{timestamp}"
        run_dir = merged_session_dir / run_id
        run_dir.mkdir(parents=True, exist_ok=True)
        final_file = run_dir / "queries.jsonl"
        
        # Save JSONL
        with open(final_file, "w", encoding="utf-8") as f:
            for q in queries:
                f.write(json.dumps(q) + "\n")
                
        # Save prompt context placeholder
        prompt_file = run_dir / "prompt.txt"
        prompt_file.write_text("Merged from multiple previous sessions. See original sessions for raw prompts.", encoding="utf-8")
        
        # Save Excel row
        save_output_metadata(
            file_path=merged_session_dir,
            MODEL_NAME=model_name,
            run_id=run_id,
            num_queries=len(queries),
            elapsed=merged_times[model_name], # Calculated proportionally from source folders
            TEMPERATURE=TEMPERATURE,
            run_dir=run_dir,
            queries_path=final_file
        )
        
    print(f"\n[success] Merge complete! You can now analyze this unified session by running the comparative metrics:")
    print(f"python metrics/main.py {merged_session_id}")
    print(f"python metrics/compare_models.py {merged_session_id}")

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Merge previous sessions into a single unified session.")
    parser.add_argument("--limit", type=int, default=None, help="Maximum number of queries per model. Defaults to TOTAL_QUERIES in .env.")
    parser.add_argument("--env", action="store_true", help="Load argument defaults from .env file")
    args = parser.parse_args()

    # If --env is passed, override defaults with .env values
    if args.env:
        from dotenv import load_dotenv
        load_dotenv()

        if args.limit is None:
            _limit = os.getenv("TOTAL_QUERIES")
            if _limit:
                args.limit = int(_limit)

        print(f"[env] Loaded defaults from .env: limit={args.limit}")
    
    merge_sessions(target_limit=args.limit)

