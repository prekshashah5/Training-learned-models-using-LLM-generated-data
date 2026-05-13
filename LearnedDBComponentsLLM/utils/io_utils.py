"""
io_utils.py
JSON, JSONL, and Excel I/O utilities.
"""

import json
import os
import tempfile
from pathlib import Path
from typing import Any, List, Optional
from datetime import datetime


def read_json_file(file_path: str) -> list:
    """
    Reads and parses a JSON or JSONL file.
    """
    with open(file_path, "r", encoding="utf-8") as f:
        content = f.read().strip()

    try:
        return json.loads(content)
    except json.JSONDecodeError:
        # Try JSON Lines format
        data = []
        decoder = json.JSONDecoder()
        idx = 0
        while idx < len(content):
            while idx < len(content) and content[idx].isspace():
                idx += 1
            if idx >= len(content):
                break
            obj, end_idx = decoder.raw_decode(content[idx:])
            if isinstance(obj, list):
                data.extend(obj)
            else:
                data.append(obj)
            idx += end_idx
        return data


def write_json_file(path, data):
    """
    Atomically write JSON to disk.
    Prevents corruption if the process crashes mid-write.
    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    with tempfile.NamedTemporaryFile(
        mode="w",
        delete=False,
        dir=path.parent,
        suffix=".tmp",
        encoding="utf-8",
    ) as tmp:
        json.dump(data, tmp, indent=2)
        tmp.flush()
        os.fsync(tmp.fileno())
        temp_name = tmp.name

    os.replace(temp_name, path)


def append_queries_to_temp(queries, temp_file: Path):
    """Append queries as JSONL to a temp file."""
    temp_file.parent.mkdir(parents=True, exist_ok=True)
    with temp_file.open("a", encoding="utf-8") as f:
        for q in queries:
            f.write(json.dumps(q, ensure_ascii=False) + "\n")


def load_queries_from_temp(temp_file: Path) -> list:
    """Load queries from a JSONL temp file."""
    queries = []
    if not temp_file.exists():
        return queries
    with temp_file.open("r", encoding="utf-8") as f:
        for line in f:
            try:
                queries.append(json.loads(line))
            except json.JSONDecodeError:
                continue
    return queries


def append_row_to_excel(path, row, header=None):
    """Append a row to an Excel file, creating it if necessary."""
    from openpyxl import Workbook, load_workbook

    if path.exists():
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        if header:
            ws.append(header)

    ws.append(row)
    wb.save(path)


def save_output_metadata(
    file_path: Path,
    MODEL_NAME: str,
    run_id: str,
    num_queries: int,
    elapsed: float,
    TEMPERATURE: float,
    run_dir: Path,
    queries_path: Path,
):
    """
    Saves metadata about a generation run into the Excel tracking file.
    """
    EXCEL_FILE = file_path / "runs.xlsx"

    append_row_to_excel(
        EXCEL_FILE,
        row=[
            datetime.now().isoformat(timespec="seconds"),
            MODEL_NAME,
            run_id,
            num_queries,
            round(elapsed, 2),
            TEMPERATURE,
            str(run_dir / "prompt.txt"),
            str(queries_path),
        ],
        header=[
            "Timestamp",
            "Model",
            "Run_ID",
            "Num_Queries",
            "Generation_Time_s",
            "Temperature",
            "Prompt_Path",
            "Queries_Path",
        ],
    )
