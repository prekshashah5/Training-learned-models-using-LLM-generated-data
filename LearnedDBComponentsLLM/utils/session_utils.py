"""
session_utils.py
Session and run directory management utilities.
"""

from pathlib import Path
from typing import Dict, List


def get_latest_json_path(output_folder: Path) -> Path:
    """Get the path to the latest queries.jsonl file in the output directory."""
    if not output_folder.exists():
        raise FileNotFoundError(f"Directory {output_folder} does not exist")

    # If the folder contains session_ directories, step into the latest one
    session_dirs = [d for d in output_folder.iterdir() if d.is_dir() and d.name.startswith("session_")]
    if session_dirs:
        output_folder = sorted(session_dirs, key=lambda d: d.name)[-1]

    run_dirs = [
        d for d in output_folder.iterdir()
        if d.is_dir() and "_run_" in d.name
    ]

    if not run_dirs:
        raise FileNotFoundError(f"No run directories found in {output_folder}")

    latest_run_dir = sorted(
        run_dirs,
        key=lambda d: d.name.split("_run_")[-1],
    )[-1]

    latest_json = latest_run_dir / "queries.jsonl"
    return latest_json


def load_model_run_data(run_dir: Path) -> dict:
    """
    Load queries and metadata from a single model run directory.

    Args:
        run_dir: Path to a run directory (e.g., output/run_20260110_023838)

    Returns:
        Dictionary with 'queries' (list), 'model_name' (str), and 'run_metadata' (dict)
    """
    from utils.io_utils import read_json_file

    queries_path = run_dir / "queries.jsonl"

    queries = []
    if queries_path.exists():
        queries = read_json_file(str(queries_path))

    # Parse model name from run_dir
    model_name_part = run_dir.name.split("_run_")[0]
    model_name = model_name_part.replace("_", ":")

    # Try getting global metadata if available
    excel_path = run_dir.parent / "runs.xlsx"
    run_metadata = {}

    if excel_path.exists():
        from openpyxl import load_workbook

        wb = load_workbook(excel_path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        run_id_idx = headers.index("Run_ID") if "Run_ID" in headers else -1

        for row in ws.iter_rows(min_row=2, values_only=True):
            if run_id_idx >= 0 and row[run_id_idx] == run_dir.name:
                run_dict = dict(zip(headers, row))
                run_metadata = {
                    "timestamp": run_dict.get("Timestamp"),
                    "model": run_dict.get("Model", model_name),
                    "run_id": run_dict.get("Run_ID"),
                    "num_queries": run_dict.get("Num_Queries", len(queries)),
                    "generation_time_s": run_dict.get("Generation_Time_s", 0),
                    "temperature": run_dict.get("Temperature"),
                }
                break

    return {
        "queries": queries,
        "model_name": model_name,
        "run_metadata": run_metadata,
        "run_dir": run_dir,
    }


def load_all_model_runs(output_folder: Path) -> tuple:
    """
    Load data from all model runs in the output folder.

    Returns:
        Tuple of (model_data dict, model_runs_metadata dict)
    """
    model_data = {}
    model_runs_metadata = {}

    run_dirs = [d for d in output_folder.iterdir() if d.is_dir() and "_run_" in d.name]

    for run_dir in run_dirs:
        run_data = load_model_run_data(run_dir)
        model_name = run_data["model_name"]

        if model_name not in model_data:
            model_data[model_name] = []
            model_runs_metadata[model_name] = []

        model_data[model_name].extend(run_data["queries"])
        model_runs_metadata[model_name].append(run_data["run_metadata"])

    return model_data, model_runs_metadata


def aggregate_model_runs_metadata(model_runs_metadata: dict) -> dict:
    """
    Aggregate metadata from multiple runs of the same model.
    """
    aggregated = {}

    for model_name, runs in model_runs_metadata.items():
        total_queries = sum(r.get("num_queries", 0) for r in runs)
        total_gen_time = sum(r.get("generation_time_s", 0) for r in runs)
        num_runs = len(runs)

        aggregated[model_name] = {
            "num_queries": total_queries,
            "generation_time_s": total_gen_time,
            "avg_generation_time_s": total_gen_time / num_runs if num_runs > 0 else 0,
            "num_runs": num_runs,
            "runs": runs,
        }

    return aggregated
