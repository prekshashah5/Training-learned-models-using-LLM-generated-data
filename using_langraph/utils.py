import re
import json
from datetime import datetime
from openpyxl import Workbook, load_workbook

def extract_json_array(text: str):
    """
    Extracts the first JSON array from text and parses it.
    Raises if not found or invalid.
    """
    match = re.search(r"\[\s*{.*?}\s*\]", text, re.DOTALL)
    if not match:
        raise ValueError("No JSON array found in model output")

    json_str = match.group(0)
    return json.loads(json_str)


def append_row_to_excel(path, row, header=None):
    if path.exists():
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        if header:
            ws.append(header)

    ws.append(row)
    wb.save(path)

def save_output_to_file(data: json, file_path: str, MODEL_NAME: str, elapsed: float, TEMPERATURE: float):
    """
    Saves data as JSON to the specified file path.
    """
    run_id = datetime.now().strftime("run_%Y%m%d_%H%M%S")
    run_dir = file_path / run_id
    run_dir.mkdir(parents=True)

    queries_path = run_dir / "queries.jsonl"
    with queries_path.open("w") as f:
        json.dump(data, f, indent=2)

    EXCEL_FILE = run_dir / "runs.xlsx"

    append_row_to_excel(
        EXCEL_FILE,
        row=[
            datetime.now().isoformat(timespec="seconds"),
            MODEL_NAME,
            run_id,
            len(data),
            round(elapsed, 2),
            TEMPERATURE,
            str(run_dir / "prompt.txt"),
            str(queries_path),
        ],
        header=[
            "Timestamp",
            "Model",
            "Run_ID",
            "Num_Queries",
            "Generation_Time_s",
            "Temperature",
            "Prompt_Path",
            "Queries_Path",
        ],
    )