
import json
import re
from typing import Any, List
from datetime import datetime
from openpyxl import Workbook, load_workbook
from pathlib import Path

def append_row_to_excel(path, row, header=None):
    if path.exists():
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        if header:
            ws.append(header)

    ws.append(row)
    wb.save(path)

def _strip_code_fences(s: str) -> str:
    t = s.strip()
    if t.startswith("```"):
        t = re.sub(r"^```(\w+)?\s*", "", t)
        t = re.sub(r"\s*```$", "", t)
    return t.strip()


def coerce_json_array(
    content: str,
    expected_len: int | None = None,
) -> List[Any]:
    """
    Robustly parse model output into a JSON array.
    - Strips ```json fences
    - Handles quoted JSON with escapes
    - Enforces array type
    - Optionally enforces expected length
    """
    if not content or not content.strip():
        raise ValueError("Empty model output")

    t = _strip_code_fences(content)

    try:
        obj = json.loads(t)
    except json.JSONDecodeError:
        # Handle quoted JSON string
        if (t.startswith('"') and t.endswith('"')) or (
            t.startswith("'") and t.endswith("'")
        ):
            inner = t[1:-1].encode("utf-8").decode("unicode_escape")
            obj = json.loads(inner)
        else:
            raise

    if not isinstance(obj, list):
        raise TypeError(f"Expected JSON array, got {type(obj)}")

    if expected_len is not None and len(obj) != expected_len:
        raise ValueError(f"Expected {expected_len} items, got {len(obj)}")

    return obj

def extract_json_array(text: str):
    """
    Extracts the first JSON array from text and parses it.
    Raises if not found or invalid.
    """
    match = re.search(r"\[\s*{.*?}\s*\]", text, re.DOTALL)
    if not match:
        raise ValueError("No JSON array found in model output")

    json_str = match.group(0)
    return json.loads(json_str)

def save_output_to_file(data: Any, file_path: str, MODEL_NAME: str, elapsed: float, TEMPERATURE: float):
    """
    Saves data as JSON to the specified file path.
    """
    run_id = datetime.now().strftime("run_%Y%m%d_%H%M%S")
    run_dir = file_path / run_id
    run_dir.mkdir(parents=True)

    queries_path = run_dir / "queries.jsonl"
    with queries_path.open("w") as f:
        json.dump(data, f, indent=2)

    EXCEL_FILE = run_dir / "runs.xlsx"

    append_row_to_excel(
        EXCEL_FILE,
        row=[
            datetime.now().isoformat(timespec="seconds"),
            MODEL_NAME,
            run_id,
            len(data),
            round(elapsed, 2),
            TEMPERATURE,
            str(run_dir / "prompt.txt"),
            str(queries_path),
        ],
        header=[
            "Timestamp",
            "Model",
            "Run_ID",
            "Num_Queries",
            "Generation_Time_s",
            "Temperature",
            "Prompt_Path",
            "Queries_Path",
        ],
    )

def read_json_file(file_path: str):
    """
    Reads and parses a JSON file.
    """
    with open(file_path, "r") as f:
        return json.load(f)
    
def get_latest_json_path(output_folder: Path) -> Path:
    # Get all run_* directories
    run_dirs = [d for d in output_folder.iterdir() if d.is_dir() and d.name.startswith("run_")]

    # Pick the most recent one by name
    latest_run_dir = sorted(run_dirs, key=lambda d: d.name)[-1]

    # Path to queries.json
    latest_json = latest_run_dir / "queries.jsonl"

    return latest_json